import time
from datetime import datetime

import random
from os.path import exists

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

# Create a Pandas dataframe from the data.
df1 = pd.DataFrame(np.reshape([10, 20, 30, 40], [1,4]), columns=['col A', 'Col B', 'Col C', 'Col D'], index=[1])
df2 = pd.DataFrame(np.reshape([101, 201, 301, 401.2342342, 102, 202.23423423, 302, 403], [8,1]), columns=['Long Table'])
df3 = pd.DataFrame(np.reshape([301, 320, 330, 340], [1,4]), columns=['col A', 'Col B', 'Col C', 'Col D'], index=[1])

# Create a Pandas Excel writer using XlsxWriter as the engine.

def example2():
    # https://xlsxwriter.readthedocs.io/working_with_pandas.html
    file_name = 'pandas_simple.xlsx'
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    sheet_name='Sheet2'

    st = 3
    df1.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=st, index=False)

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    worksheet.write(0, 0, "this is s comment string !")
    worksheet.write("A2", "this is s comment string for A2")

   # worksheet.cell(row=1,column=1).value = "This is an added comments"

    st += df1.shape[0] + 2
    df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=st, index=False)
    worksheet.conditional_format(7, 0, 14, 0,
                                 {'type': '3_color_scale'})

    numberFromat = workbook.add_format({'num_format': '#,##0.00'})
    worksheet.set_column(first_col=0, last_col=1,  cell_format=numberFromat)

    st += df2.shape[0] + 2
    df3.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=st, index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()

    print("all Done!")

def example3():
    import pandas as pd

    # Create a test df
    df = pd.DataFrame({'Name': ['Tesla', 'Tesla', 'Toyota', 'Ford', 'Ford', 'Ford'],
                       'Type': ['Model X', 'Model Y', 'Corolla', 'Bronco', 'Fiesta', 'Mustang']})

    # Create the list where we 'll capture the cells that appear for 1st time,
    # add the 1st row and we start checking from 2nd row until end of df
    startCells = [1]
    for row in range(2, len(df) + 1):
        if (df.loc[row - 1, 'Name'] != df.loc[row - 2, 'Name']):
            startCells.append(row)

    writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter',merge_cells=False)
    df.to_excel(writer, sheet_name='Example 3', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Example 3']
    merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})

    lastRow = len(df)

    for row in startCells:
        try:
            endRow = startCells[startCells.index(row) + 1] - 1
            if row == endRow:
                worksheet.write(row, 0, df.loc[row - 1, 'Name'], merge_format)
            else:
                worksheet.merge_range(row, 0, endRow, 0, df.loc[row - 1, 'Name'], merge_format)
        except IndexError:
            if row == lastRow:
                worksheet.write(row, 0, df.loc[row - 1, 'Name'], merge_format)
            else:
                worksheet.merge_range(row, 0, lastRow, 0, df.loc[row - 1, 'Name'], merge_format)

    writer.save()


def cformat():
    workbook = xlsxwriter.Workbook('conditional_format.xlsx')
    worksheet1 = workbook.add_worksheet()
    worksheet2 = workbook.add_worksheet()
    worksheet3 = workbook.add_worksheet()
    worksheet4 = workbook.add_worksheet()
    worksheet5 = workbook.add_worksheet()
    worksheet6 = workbook.add_worksheet()
    worksheet7 = workbook.add_worksheet()
    worksheet8 = workbook.add_worksheet()
    worksheet9 = workbook.add_worksheet()

    # Add a format. Light red fill with dark red text.
    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006'})

    # Add a format. Green fill with dark green text.
    format2 = workbook.add_format({'bg_color': '#C6EFCE',
                                   'font_color': '#006100'})

    # Some sample data to run the conditional formatting against.
    data = [
        [34, 72, 38, 30, 75, 48, 75, 66, 84, 86],
        [6, 24, 1, 84, 54, 62, 60, 3, 26, 59],
        [28, 79, 97, 13, 85, 93, 93, 22, 5, 14],
        [27, 71, 40, 17, 18, 79, 90, 93, 29, 47],
        [88, 25, 33, 23, 67, 1, 59, 79, 47, 36],
        [24, 100, 20, 88, 29, 33, 38, 54, 54, 88],
        [6, 57, 88, 28, 10, 26, 37, 7, 41, 48],
        [52, 78, 1, 96, 26, 45, 47, 33, 96, 36],
        [60, 54, 81, 66, 81, 90, 80, 93, 12, 55],
        [70, 5, 46, 14, 71, 19, 66, 36, 41, 21],
    ]

    ###############################################################################
    #
    # Example 1.
    #
    caption = ('Cells with values >= 50 are in light red. '
               'Values < 50 are in light green.')

    # Write the data.
    worksheet1.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet1.write_row(row + 2, 1, row_data)

    # Write a conditional format over a range.
    worksheet1.conditional_format('B3:K12', {'type': 'cell',
                                             'criteria': '>=',
                                             'value': 50,
                                             'format': format1})

    # Write another conditional format over the same range.
    worksheet1.conditional_format('B3:K12', {'type': 'cell',
                                             'criteria': '<',
                                             'value': 50,
                                             'format': format2})

    ###############################################################################
    #
    # Example 2.
    #
    caption = ('Values between 30 and 70 are in light red. '
               'Values outside that range are in light green.')

    worksheet2.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet2.write_row(row + 2, 1, row_data)

    worksheet2.conditional_format('B3:K12', {'type': 'cell',
                                             'criteria': 'between',
                                             'minimum': 30,
                                             'maximum': 70,
                                             'format': format1})

    worksheet2.conditional_format('B3:K12', {'type': 'cell',
                                             'criteria': 'not between',
                                             'minimum': 30,
                                             'maximum': 70,
                                             'format': format2})

    ###############################################################################
    #
    # Example 3.
    #
    caption = ('Duplicate values are in light red. '
               'Unique values are in light green.')

    worksheet3.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet3.write_row(row + 2, 1, row_data)

    worksheet3.conditional_format('B3:K12', {'type': 'duplicate',
                                             'format': format1})

    worksheet3.conditional_format('B3:K12', {'type': 'unique',
                                             'format': format2})

    ###############################################################################
    #
    # Example 4.
    #
    caption = ('Above average values are in light red. '
               'Below average values are in light green.')

    worksheet4.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet4.write_row(row + 2, 1, row_data)

    worksheet4.conditional_format('B3:K12', {'type': 'average',
                                             'criteria': 'above',
                                             'format': format1})

    worksheet4.conditional_format('B3:K12', {'type': 'average',
                                             'criteria': 'below',
                                             'format': format2})

    ###############################################################################
    #
    # Example 5.
    #
    caption = ('Top 10 values are in light red. '
               'Bottom 10 values are in light green.')

    worksheet5.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet5.write_row(row + 2, 1, row_data)

    worksheet5.conditional_format('B3:K12', {'type': 'top',
                                             'value': '10',
                                             'format': format1})

    worksheet5.conditional_format('B3:K12', {'type': 'bottom',
                                             'value': '10',
                                             'format': format2})

    ###############################################################################
    #
    # Example 6.
    #
    caption = ('Cells with values >= 50 are in light red. '
               'Values < 50 are in light green. Non-contiguous ranges.')

    # Write the data.
    worksheet6.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet6.write_row(row + 2, 1, row_data)

    # Write a conditional format over a range.
    worksheet6.conditional_format('B3:K6', {'type': 'cell',
                                            'criteria': '>=',
                                            'value': 50,
                                            'format': format1,
                                            'multi_range': 'B3:K6 B9:K12'})

    # Write another conditional format over the same range.
    worksheet6.conditional_format('B3:K6', {'type': 'cell',
                                            'criteria': '<',
                                            'value': 50,
                                            'format': format2,
                                            'multi_range': 'B3:K6 B9:K12'})

    ###############################################################################
    #
    # Example 7.
    #
    caption = 'Examples of color scales with default and user colors.'

    data = range(1, 13)

    worksheet7.write('A1', caption)

    worksheet7.write('B2', "2 Color Scale")
    worksheet7.write('D2', "2 Color Scale + user colors")

    worksheet7.write('G2', "3 Color Scale")
    worksheet7.write('I2', "3 Color Scale + user colors")

    for row, row_data in enumerate(data):
        worksheet7.write(row + 2, 1, row_data)
        worksheet7.write(row + 2, 3, row_data)
        worksheet7.write(row + 2, 6, row_data)
        worksheet7.write(row + 2, 8, row_data)

    worksheet7.conditional_format('B3:B14', {'type': '2_color_scale'})

    worksheet7.conditional_format('D3:D14', {'type': '2_color_scale',
                                             'min_color': "#FF0000",
                                             'max_color': "#00FF00"})

    worksheet7.conditional_format('G3:G14', {'type': '3_color_scale'})

    worksheet7.conditional_format('I3:I14', {'type': '3_color_scale',
                                             'min_color': "#C5D9F1",
                                             'mid_color': "#8DB4E3",
                                             'max_color': "#538ED5"})

    ###############################################################################
    #
    # Example 8.
    #
    caption = 'Examples of data bars.'

    worksheet8.write('A1', caption)

    worksheet8.write('B2', "Default data bars")
    worksheet8.write('D2', "Bars only")
    worksheet8.write('F2', "With user color")
    worksheet8.write('H2', "Solid bars")
    worksheet8.write('J2', "Right to left")
    worksheet8.write('L2', "Excel 2010 style")
    worksheet8.write('N2', "Negative same as positive")

    data = range(1, 13)
    for row, row_data in enumerate(data):
        worksheet8.write(row + 2, 1, row_data)
        worksheet8.write(row + 2, 3, row_data)
        worksheet8.write(row + 2, 5, row_data)
        worksheet8.write(row + 2, 7, row_data)
        worksheet8.write(row + 2, 9, row_data)

    data = [-1, -2, -3, -2, -1, 0, 1, 2, 3, 2, 1, 0]
    for row, row_data in enumerate(data):
        worksheet8.write(row + 2, 11, row_data)
        worksheet8.write(row + 2, 13, row_data)

    worksheet8.conditional_format('B3:B14', {'type': 'data_bar'})

    worksheet8.conditional_format('D3:D14', {'type': 'data_bar',
                                             'bar_only': True})

    worksheet8.conditional_format('F3:F14', {'type': 'data_bar',
                                             'bar_color': '#63C384'})

    worksheet8.conditional_format('H3:H14', {'type': 'data_bar',
                                             'bar_solid': True})

    worksheet8.conditional_format('J3:J14', {'type': 'data_bar',
                                             'bar_direction': 'right'})

    worksheet8.conditional_format('L3:L14', {'type': 'data_bar',
                                             'data_bar_2010': True})

    worksheet8.conditional_format('N3:N14', {'type': 'data_bar',
                                             'bar_negative_color_same': True,
                                             'bar_negative_border_color_same': True})

    ###############################################################################
    #
    # Example 9.
    #
    caption = 'Examples of conditional formats with icon sets.'

    data = [
        [1, 2, 3],
        [1, 2, 3],
        [1, 2, 3],
        [1, 2, 3],
        [1, 2, 3, 4],
        [1, 2, 3, 4, 5],
        [1, 2, 3, 4, 5],
    ]

    worksheet9.write('A1', caption)

    for row, row_data in enumerate(data):
        worksheet9.write_row(row + 2, 1, row_data)

    worksheet9.conditional_format('B3:D3', {'type': 'icon_set',
                                            'icon_style': '3_traffic_lights'})

    worksheet9.conditional_format('B4:D4', {'type': 'icon_set',
                                            'icon_style': '3_traffic_lights',
                                            'reverse_icons': True})

    worksheet9.conditional_format('B5:D5', {'type': 'icon_set',
                                            'icon_style': '3_traffic_lights',
                                            'icons_only': True})

    worksheet9.conditional_format('B6:D6', {'type': 'icon_set',
                                            'icon_style': '3_arrows'})

    worksheet9.conditional_format('B7:E7', {'type': 'icon_set',
                                            'icon_style': '4_arrows'})

    worksheet9.conditional_format('B8:F8', {'type': 'icon_set',
                                            'icon_style': '5_arrows'})

    worksheet9.conditional_format('B9:F9', {'type': 'icon_set',
                                            'icon_style': '5_ratings'})

    workbook.close()



example3()